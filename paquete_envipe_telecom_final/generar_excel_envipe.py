#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Lee los TMod_Vic.dbf (o ZIP) de ENVIPE y genera UN SOLO Excel con el
desglose de tipo de delito x modalidad/medio de comisión, listo para
análisis (sin columnas técnicas de la encuesta como "pregunta" o
"variable").

Reutiliza el parser de DBF de procesar_envipe_dbf_desglose_encuesta.py
(discover_sources, open_source, iter_selected_records, process_records,
build_output_rows) en lugar de reescribirlo: ese módulo ya resuelve las
diferencias de catálogo entre ediciones (2011-2025) y el factor de
expansión FAC_DEL. Este script sólo se encarga de:

    1) recolectar las filas crudas para el rango de ediciones pedido,
    2) reclasificarlas en categorías útiles (totales, fraude, extorsión
       por tipo, extorsión por lugar de contacto, medio de comisión),
    3) escribir un .xlsx con una pestaña por categoría + una pestaña
       "Desglose completo" + una pestaña "Notas" con la metodología.

Requisito: procesar_envipe_dbf_desglose_encuesta.py debe estar en la
misma carpeta que este archivo (se importa como módulo).

Uso:
    python generar_excel_envipe.py --dir conjunto_de_datos \
        --salida ENVIPE_DESGLOSE_ENCUESTA.xlsx

    python generar_excel_envipe.py --dir conjunto_de_datos \
        --start-edition 2016 --end-edition 2025
"""

from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "Falta la librería openpyxl. Instálala con:\n"
        "    pip install openpyxl"
    ) from exc

# --------------------------------------------------------------------------
# Reutiliza el parser de DBF ya existente en lugar de duplicarlo.
# --------------------------------------------------------------------------
sys.path.insert(0, str(Path(__file__).resolve().parent))
try:
    import procesar_envipe_dbf_desglose_encuesta as envipe_dbf
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "No se encontró procesar_envipe_dbf_desglose_encuesta.py.\n"
        "Coloca generar_excel_envipe.py en la misma carpeta que ese archivo."
    ) from exc


MIN_EDITION = envipe_dbf.MIN_EDITION
MAX_EDITION = envipe_dbf.MAX_EDITION
DEFAULT_XLSX = "ENVIPE_DESGLOSE_ENCUESTA.xlsx"

TOTAL_QUESTION = "Total por tipo de delito"

# Campos "tidy" finales, sin id/pregunta/variable/codigo_respuesta/tipo_respuesta.
FULL_FIELDS = [
    "anio",
    "categoria",
    "delito",
    "modalidad_comision",
    "estimacion",
    "muestra",
    "porcentaje",
    "base_referencia",
]
GROUP_FIELDS = [
    "anio",
    "delito",
    "modalidad_comision",
    "estimacion",
    "muestra",
    "porcentaje",
    "base_referencia",
]
TOTALS_FIELDS = ["anio", "delito", "estimacion", "muestra"]

HEADER_LABELS = {
    "anio": "Año",
    "categoria": "Categoría",
    "delito": "Delito",
    "modalidad_comision": "Modalidad / medio de comisión",
    "estimacion": "Estimación (personas o casos)",
    "muestra": "Muestra (observaciones)",
    "porcentaje": "Porcentaje",
    "base_referencia": "Base de referencia del porcentaje",
}

CATEGORY_LABELS = {
    "totales": "Total por tipo de delito",
    "fraude": "Fraude - modalidad",
    "extorsion_tipo": "Extorsión - tipo",
    "extorsion_lugar": "Extorsión - lugar de contacto presencial",
    "medio_comision": "Medio de comisión por delito",
    "otros": "Otros",
}

# (clave interna, nombre de pestaña, columnas)
SHEET_GROUPS: list[tuple[str, str, list[str]]] = [
    ("totales", "Totales por delito", TOTALS_FIELDS),
    ("fraude", "Fraude por modalidad", GROUP_FIELDS),
    ("extorsion_tipo", "Extorsión por tipo", GROUP_FIELDS),
    ("extorsion_lugar", "Extorsión - lugar contacto", GROUP_FIELDS),
    ("medio_comision", "Medio de comisión", GROUP_FIELDS),
]


# --------------------------------------------------------------------------
# Paso 1: recolectar filas crudas (mismo formato que el CSV original)
# --------------------------------------------------------------------------
def build_raw_rows(
    root: Path,
    start_edition: int,
    end_edition: int,
    require_all: bool,
) -> list[dict[str, object]]:
    """Recorre los DBF/ZIP de ENVIPE y devuelve las filas crudas, en el
    mismo formato que produce procesar_envipe_dbf_desglose_encuesta.py."""

    sources = envipe_dbf.discover_sources(root, start_edition, end_edition)
    if not sources:
        raise SystemExit("No se encontraron archivos TMod_Vic.dbf.")

    found = {source.edition for source in sources}
    expected = set(range(start_edition, end_edition + 1))
    missing = sorted(expected - found)
    if missing:
        message = "Faltan ediciones: " + ", ".join(map(str, missing))
        if require_all:
            raise SystemExit(message)
        logging.warning(message)

    raw_rows: list[dict[str, object]] = []
    for source in sources:
        logging.info(
            "Procesando ENVIPE %s: %s", source.edition, source.display_name
        )
        try:
            with envipe_dbf.open_source(source) as fh:
                records = envipe_dbf.iter_selected_records(
                    fh, envipe_dbf.required_fields(source.edition)
                )
                processed = envipe_dbf.process_records(source.edition, records)
        except Exception as exc:
            raise SystemExit(f"Error en ENVIPE {source.edition}: {exc}") from exc

        invalid = processed[-1]
        if invalid:
            logging.warning(
                "ENVIPE %s: se excluyeron %s registros con FAC_DEL inválido",
                source.edition,
                invalid,
            )
        raw_rows.extend(envipe_dbf.build_output_rows(source, processed))

    section_order = {
        "I. Todos los tipos de delito": 1,
        "IV. Fraude": 4,
        "V. Extorsión": 5,
    }
    raw_rows.sort(
        key=lambda row: (
            int(row["anio"]),
            section_order.get(str(row["seccion"]), 99),
            str(row["codigo_delito"]),
            str(row["pregunta"]),
            str(row["codigo_respuesta"]),
        )
    )
    return raw_rows


# --------------------------------------------------------------------------
# Paso 2: reclasificar cada fila cruda en una categoría útil ("tidy")
# --------------------------------------------------------------------------
def _to_float(value: object) -> float | None:
    if value in (None, ""):
        return None
    return float(value)  # type: ignore[arg-type]


def classify(row: dict[str, object]) -> str:
    """Determina a qué categoría/pestaña pertenece una fila cruda, a partir
    de la sección y la pregunta (no sólo la sección: la Sección I incluye,
    desde ENVIPE 2025, un desglose adicional de medio de comisión)."""
    pregunta = str(row["pregunta"])
    seccion = str(row["seccion"])

    if pregunta == TOTAL_QUESTION:
        return "totales"
    if seccion == "IV. Fraude":
        return "fraude"
    if seccion == "V. Extorsión":
        return "extorsion_lugar" if pregunta.startswith("5.1a") else "extorsion_tipo"
    if pregunta.startswith("1.5a"):
        return "medio_comision"
    return "otros"


def tidy_row(row: dict[str, object]) -> dict[str, object]:
    kind = classify(row)

    if kind == "totales":
        modalidad: object = "Total"
        porcentaje = 100.0
    else:
        modalidad = row["respuesta"]
        porcentaje = _to_float(row["porcentaje_base"])

    return {
        "anio": int(row["anio"]),
        "categoria": CATEGORY_LABELS[kind],
        "delito": row["delito"],
        "modalidad_comision": modalidad,
        "estimacion": int(row["estimacion"]),
        "muestra": int(row["muestra"]),
        "porcentaje": porcentaje,
        "base_referencia": row["base_porcentaje"],
        "_kind": kind,
    }


def build_tidy_rows(raw_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    return [tidy_row(row) for row in raw_rows]


# --------------------------------------------------------------------------
# Paso 3: escribir el Excel
# --------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E5F", end_color="1F4E5F", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF")
BASE_FONT = Font(name="Calibri")


def write_sheet(ws: Worksheet, rows: list[dict[str, object]], fields: list[str]) -> None:
    ws.append([HEADER_LABELS[f] for f in fields])
    for col_idx in range(1, len(fields) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    for row in rows:
        values = []
        for f in fields:
            v = row[f]
            if f == "porcentaje" and v is not None:
                v = v / 100  # se guarda como fracción para usar formato % nativo
            values.append(v)
        ws.append(values)

    last_row = ws.max_row
    for col_idx, f in enumerate(fields, start=1):
        col_letter = get_column_letter(col_idx)
        if f in ("estimacion", "muestra"):
            for r in range(2, last_row + 1):
                ws.cell(row=r, column=col_idx).number_format = "#,##0"
        elif f == "porcentaje":
            for r in range(2, last_row + 1):
                ws.cell(row=r, column=col_idx).number_format = "0.00%"

        max_len = max(
            [len(HEADER_LABELS[f])] + [len(str(row[f])) for row in rows] + [8]
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)

    for r in range(2, last_row + 1):
        for c in range(1, len(fields) + 1):
            ws.cell(row=r, column=c).font = BASE_FONT


NOTES: list[tuple[str, str]] = [
    ("Fuente", "Microdatos TMod_Vic.dbf de ENVIPE (INEGI), ediciones {start}-{end}."),
    (
        "Convención de año",
        "'Año' es el año de referencia de los hechos (la edición ENVIPE N "
        "corresponde al año N-1, ej. ENVIPE 2025 = año de referencia 2024).",
    ),
    (
        "estimacion",
        "Número estimado de víctimas/casos, expandido con el factor FAC_DEL "
        "y redondeado al entero más cercano.",
    ),
    (
        "muestra",
        "Número de observaciones (entrevistas) que sostienen la estimación, "
        "SIN expandir.",
    ),
    (
        "porcentaje",
        "Porcentaje sobre la base indicada en 'base_referencia' (no sobre "
        "el total general de delitos).",
    ),
    (
        "Cambio de metodología en ENVIPE 2025 (año de referencia 2024)",
        "Extorsión separa 'tipo' (ya no incluye 'telefónica' como opción "
        "propia) de 'lugar de contacto presencial'. Además se agrega la "
        "pregunta de medio de comisión (internet, llamada telefónica, "
        "contacto presencial, otro) para fraude bancario, fraude al "
        "consumidor, extorsión, amenazas y hostigamiento sexual. No "
        "compares estas series directamente contra años anteriores sin "
        "considerar este cambio.",
    ),
    (
        "Delitos sin modalidad",
        "Robo, vandalismo, lesiones, secuestro y violación no tienen "
        "desglose por modalidad en el DBF; sólo aparecen en la pestaña "
        "'Totales por delito'.",
    ),
]


def write_notes_sheet(ws: Worksheet, start_edition: int, end_edition: int) -> None:
    ws.append(["Nota", "Detalle"])
    for col_idx in (1, 2):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for label, text in NOTES:
        ws.append([label, text.format(start=start_edition, end=end_edition)])
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 100
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=1).alignment = Alignment(vertical="top")


def generate_workbook(
    tidy_rows: list[dict[str, object]],
    start_edition: int,
    end_edition: int,
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    sheet_all = wb.create_sheet("Desglose completo")
    write_sheet(sheet_all, tidy_rows, FULL_FIELDS)

    for kind, sheet_name, fields in SHEET_GROUPS:
        subset = [r for r in tidy_rows if r["_kind"] == kind]
        if not subset:
            continue
        ws = wb.create_sheet(sheet_name)
        write_sheet(ws, subset, fields)

    write_notes_sheet(wb.create_sheet("Notas"), start_edition, end_edition)
    return wb


# --------------------------------------------------------------------------
def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Lee los DBF/ZIP de ENVIPE (TMod_Vic) y genera un Excel con el "
            "desglose de tipo de delito x modalidad de comisión."
        )
    )
    parser.add_argument(
        "--dir",
        type=Path,
        default=Path("conjunto_de_datos"),
        help="Directorio con DBF o ZIP de ENVIPE.",
    )
    parser.add_argument(
        "--salida",
        type=Path,
        default=Path(DEFAULT_XLSX),
        help=f"Excel de salida. Predeterminado: {DEFAULT_XLSX}",
    )
    parser.add_argument("--start-edition", type=int, default=MIN_EDITION)
    parser.add_argument("--end-edition", type=int, default=MAX_EDITION)
    parser.add_argument(
        "--require-all",
        action="store_true",
        help="Detiene el proceso si falta alguna edición.",
    )
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    if not args.dir.is_dir():
        raise SystemExit(f"No existe el directorio: {args.dir}")

    raw_rows = build_raw_rows(
        args.dir, args.start_edition, args.end_edition, args.require_all
    )
    tidy_rows = build_tidy_rows(raw_rows)

    wb = generate_workbook(tidy_rows, args.start_edition, args.end_edition)

    args.salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.salida)
    print(f"Archivo generado: {args.salida}")
    print(f"Filas procesadas: {len(tidy_rows)}")


if __name__ == "__main__":
    main()
